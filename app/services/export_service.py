from __future__ import annotations

import csv
import io
from collections.abc import Iterator
from datetime import datetime

from fastapi.responses import Response, StreamingResponse
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.models import Document, DocumentReference
from app.services.analytics_service import (
    get_dashboard_summary,
    get_domain_summary,
    get_error_summary,
)
from app.services.results_service import get_reference_summary, iter_references


def export_csv(
    session: Session,
    filters: dict,
) -> StreamingResponse:
    def generate() -> Iterator[str]:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(
            [
                "document_id",
                "filename",
                "page_number",
                "reference_class",
                "source_type",
                "raw_reference",
                "final_url",
                "resolution_status",
            ]
        )
        yield buffer.getvalue()
        buffer.seek(0)
        buffer.truncate(0)

        for row in iter_references(session, **filters):
            writer.writerow(
                [
                    row.document_id,
                    row.filename,
                    row.page_number,
                    row.reference_class,
                    row.source_type,
                    row.raw_reference,
                    row.final_url or "",
                    row.resolution_status,
                ]
            )
            yield buffer.getvalue()
            buffer.seek(0)
            buffer.truncate(0)

    return StreamingResponse(
        generate(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="olre-results.csv"'},
    )


def export_markdown(
    session: Session,
    filters: dict,
) -> StreamingResponse:
    summary = get_reference_summary(session, **filters)

    def generate() -> Iterator[str]:
        yield "# OLRE Extraction Report\n\n"
        yield "## Summary\n"
        yield f"- Total documents: {summary['total_documents']}\n"
        yield f"- Total references: {summary['total_references']}\n"
        yield f"- Resolved: {summary['resolved']}\n"
        yield f"- Failed: {summary['failed']}\n\n"
        yield "## Details\n\n"

        current_filename: str | None = None
        for row in iter_references(session, **filters):
            if row.filename != current_filename:
                if current_filename is not None:
                    yield "\n---\n\n"
                current_filename = row.filename
                yield f"### File: {row.filename}\n\n"
                yield "| Page | Type | Raw | Final | Status |\n"
                yield "|------|------|-----|-------|--------|\n"

            raw_reference = (row.raw_reference or "").replace("|", "\\|")
            final_url = (row.final_url or "").replace("|", "\\|")
            yield (
                f"| {row.page_number} | {row.reference_class} | {raw_reference} | "
                f"{final_url} | {row.resolution_status} |\n"
            )

        if current_filename is None:
            yield "_No references found._\n"

    return StreamingResponse(
        generate(),
        media_type="text/markdown; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="olre-report.md"'},
    )


def _style_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.style = "Headline 4"
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)


def export_excel(session: Session, filters: dict) -> Response:
    from openpyxl import Workbook

    _ = filters
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    summary = get_dashboard_summary(session)
    summary_sheet.append(["metric", "value"])
    for key in [
        "total_documents",
        "total_references",
        "processed_documents",
        "failed_documents",
        "duplicate_documents",
        "resolved_urls",
        "failed_urls",
        "broken_link_rate",
        "qr_count",
        "text_count",
        "ocr_count",
    ]:
        summary_sheet.append([key, summary[key]])
    _style_sheet(summary_sheet)

    documents_sheet = workbook.create_sheet("Documents")
    documents_sheet.append(
        [
            "document_id",
            "original_file_name",
            "processing_status",
            "processing_error_type",
            "processing_error_detail",
            "page_count",
            "processed_at",
        ]
    )
    for row in session.execute(
        select(
            Document.id,
            Document.original_file_name,
            Document.processing_status,
            Document.processing_error_type,
            Document.processing_error_detail,
            Document.page_count,
            Document.processed_at,
        ).order_by(Document.id.asc())
    ):
        documents_sheet.append(
            [
                row.id,
                row.original_file_name,
                row.processing_status,
                row.processing_error_type,
                row.processing_error_detail,
                row.page_count,
                row.processed_at.isoformat() if row.processed_at else "",
            ]
        )
    _style_sheet(documents_sheet)

    references_sheet = workbook.create_sheet("References")
    references_sheet.append(
        [
            "reference_id",
            "document_id",
            "original_file_name",
            "source_type",
            "page_number",
            "raw_reference",
            "resolved_url",
            "resolution_status",
            "resolution_error_type",
            "resolution_error_detail",
        ]
    )
    reference_statement = (
        select(
            DocumentReference.id,
            DocumentReference.document_id,
            Document.original_file_name,
            DocumentReference.source_type,
            DocumentReference.page_number,
            DocumentReference.raw_reference,
            DocumentReference.final_url,
            DocumentReference.resolution_status,
            DocumentReference.resolution_error_type,
            DocumentReference.resolution_error_detail,
        )
        .select_from(DocumentReference)
        .join(Document, DocumentReference.document_id == Document.id)
        .order_by(DocumentReference.id.asc())
    )
    for row in session.execute(reference_statement):
        references_sheet.append(list(row))
    _style_sheet(references_sheet)

    domains_sheet = workbook.create_sheet("Domains")
    domains_sheet.append(["domain", "total_references", "resolved_count", "failed_count", "success_rate"])
    for row in get_domain_summary(session, limit=1000):
        domains_sheet.append(
            [row.domain, row.total_references, row.resolved_count, row.failed_count, row.success_rate]
        )
    _style_sheet(domains_sheet)

    errors_sheet = workbook.create_sheet("Errors")
    errors_sheet.append(["error_scope", "error_type", "count"])
    errors = get_error_summary(session, limit=1000)
    for row in errors["processing_errors"]:
        errors_sheet.append(["document", row["error_type"], row["count"]])
    for row in errors["resolution_errors"]:
        errors_sheet.append(["reference", row["error_type"], row["count"]])
    _style_sheet(errors_sheet)

    output = io.BytesIO()
    workbook.save(output)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    filename = f"olre_report_{timestamp}.xlsx"
    return Response(
        output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
